# Copyright (c) 2022-2026 MKM Research Labs.
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Generate a filled demo upload workbook (docs/bulkupload/cdm_demo.xlsx).

Five records each of Portfolio (property), Commercial, Mortgage and Commercial
Loan, ready to bulk-upload into the CDM Asset Review tool to demo the feature.
Values are pulled from the real thames sandbox data and re-validated through the
same ``cdm_edit`` rules the importer uses, so every row imports cleanly; ids are
re-stamped as DEMO-* so they don't collide with the seeded portfolio. Sheet names
match the template's so the tool's Upload reads them directly.

Run:  python tools/cdm_property_editor/cdm_demo.py
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

TOOL_DIR = Path(__file__).resolve().parent
REPO_ROOT = TOOL_DIR.parent.parent
SRC_DIR = REPO_ROOT / "src"
for _p in (str(SRC_DIR), str(REPO_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)  # pragma: no cover (entrypoint path bootstrap)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from cdm_edit import descriptor_at, validate_value  # noqa: E402
from port.cdm.asset.commercial.schema import COMMERCIAL_SCHEMA  # noqa: E402
from port.cdm.asset.loan.schema import MORTGAGE_SCHEMA  # noqa: E402
from port.cdm.asset.residential.schema import PROPERTY_SCHEMA  # noqa: E402

INPUT_DIR = REPO_ROOT / "data" / "input" / "thames"
OUT_PATH = REPO_ROOT / "docs" / "bulkupload" / "cdm_demo.xlsx"
N = 5

_PH, _CA, _RL, _ML = "PropertyHeader", "CommercialAsset", "RLoan", "Mortgage"

# (sheet, schema, data file, container, id path, id prefix, curated columns)
ASSETS = [
    ("Portfolio (Template)", PROPERTY_SCHEMA, "property.json", "properties",
     f"{_PH}.Header.PropertyID", "DEMO-PROP", [
         f"{_PH}.Header.PropertyID", f"{_PH}.Location.BuildingNumber",
         f"{_PH}.Location.StreetName", f"{_PH}.Location.TownCity",
         f"{_PH}.Location.Postcode", f"{_PH}.Header.propertyType",
         f"{_PH}.Valuation.PropertyValue", f"{_PH}.Location.LatitudeDegrees",
         f"{_PH}.Location.LongitudeDegrees", f"{_PH}.Construction.FloorLevelMeters",
         f"{_PH}.RiskAssessment.GroundLevelMeters"]),
    ("Commercial (Template)", COMMERCIAL_SCHEMA, "commercial.json", "commercial_assets",
     f"{_CA}.Header.PropertyID", "DEMO-CPROP", [
         f"{_CA}.Header.PropertyID", f"{_CA}.Location.BuildingNumber",
         f"{_CA}.Location.StreetName", f"{_CA}.Location.TownCity",
         f"{_CA}.Location.Postcode", f"{_CA}.CommercialAttributes.CommercialType",
         f"{_CA}.Valuation.PropertyValue", f"{_CA}.Location.LatitudeDegrees",
         f"{_CA}.Location.LongitudeDegrees", f"{_CA}.Construction.FloorLevelMeters",
         f"{_CA}.RiskAssessment.GroundLevelMeters"]),
    ("Mortgage (Template)", MORTGAGE_SCHEMA, "loan.json", "loans",
     f"{_RL}.Header.RLoanID", "DEMO-RLOAN", [
         f"{_RL}.Header.RLoanID", f"{_RL}.Header.PropertyID",
         f"{_RL}.FinancialTerms.OriginalLoan", f"{_RL}.CurrentStatus.OutstandingBalance",
         f"{_RL}.CurrentStatus.AccountStatus", f"{_RL}.Features.MortgageType",
         f"{_RL}.Features.RepaymentType", f"{_RL}.FinancialTerms.OriginalRateType"]),
    ("Commercial Loan (Template)", {"Mortgage": MORTGAGE_SCHEMA["RLoan"]},
     "commercial_loan.json", "commercial_loans",
     f"{_ML}.Header.RLoanID", "DEMO-CLOAN", [
         f"{_ML}.Header.RLoanID", f"{_ML}.Header.PropertyID",
         f"{_ML}.FinancialTerms.OriginalLoan", f"{_ML}.CurrentStatus.OutstandingBalance",
         f"{_ML}.CurrentStatus.AccountStatus", f"{_ML}.Features.MortgageType"]),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
WRAP = Alignment(vertical="top", wrap_text=True)


def _get(rec: dict, path: str):
    node = rec
    for seg in path.split("."):
        if not isinstance(node, dict) or seg not in node:
            return None
        node = node[seg]
    return node


def _demo_rows(schema, records, id_path, prefix, columns):
    """Up to N validated rows {path: value}; id re-stamped DEMO-*."""
    rows = []
    for i, rec in enumerate(records[:N], start=1):
        row = {id_path: f"{prefix}-{i:02d}"}
        for path in columns:
            if path == id_path:
                continue
            val = _get(rec, path)
            if val in (None, ""):
                continue
            desc = descriptor_at(schema, path)
            if desc is None:
                continue
            ok, coerced, _ = validate_value(path, desc, val)
            if ok:
                row[path] = coerced
        rows.append(row)
    return rows


def _sheet(wb, title, columns, rows):
    ws = wb.create_sheet(title)
    # keep only columns that at least one row populates, in curated order
    used = [c for c in columns if any(c in r for r in rows)]
    ws.append(used)
    for col in range(1, len(used) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill, cell.font, cell.alignment = HEADER_FILL, HEADER_FONT, WRAP
        ws.column_dimensions[get_column_letter(col)].width = 22
    for r in rows:
        ws.append([r.get(c, "") for c in used])
    ws.freeze_panes = "A2"


def build() -> Path:
    wb = Workbook()
    wb.remove(wb.active)
    for title, schema, fname, container, id_path, prefix, columns in ASSETS:
        data = json.loads((INPUT_DIR / fname).read_text())
        records = data.get(container, []) if isinstance(data, dict) else data
        rows = _demo_rows(schema, records, id_path, prefix, columns)
        _sheet(wb, title, columns, rows)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
