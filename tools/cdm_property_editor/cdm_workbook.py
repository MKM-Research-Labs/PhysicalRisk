# Copyright (c) 2022-2026 MKM Research Labs. All rights reserved.

# This software is licensed by MKM Research Labs for non-commercial 
# research and educational use only. Any commercial use, including 
# but not limited to use in or for products or services offered for sale, 
# internal business operations intended for commercial advantage, or
# research and development conducted for a commercial entity, is expressly
# prohibited unless separately authorized in writing by MKM Research Labs.

# Use, reproduction, distribution, or modification of this code is subject to the
# terms and conditions of the license agreement provided with this software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Generate the CDM upload workbook (docs/bulkupload/cdm_upload_workbook.xlsx).

One workbook, four asset classes (Portfolio / Mortgage / Commercial /
Commercial Loan), each with two sheets:

  * ``<Asset> (Dictionary)`` — one row per CDM leaf field: path, section,
    field, type, allowed values, description. The field spec / contract.
  * ``<Asset> (Template)``  — the same fields as COLUMNS (CDM dotted path in
    the header, with the rules as a cell comment + a dropdown for menu fields),
    blank rows beneath for data entry. The sheet a user fills in to upload.

The field set is derived from the canonical CDM Python schemas via
``cdm_edit.schema_specs`` — the same flattener the editor uses — so the
workbook never drifts from the schema. Run:

    source .venv/bin/activate
    python tools/cdm_property_editor/cdm_workbook.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

# --- make the CDM schemas + cdm_edit importable (mirror app.py) -------------
TOOL_DIR = Path(__file__).resolve().parent
REPO_ROOT = TOOL_DIR.parent.parent
SRC_DIR = REPO_ROOT / "src"
for _p in (str(SRC_DIR), str(REPO_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

from openpyxl import Workbook  # noqa: E402
from openpyxl.comments import Comment  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402
from openpyxl.worksheet.datavalidation import DataValidation  # noqa: E402

from cdm_edit import schema_specs  # noqa: E402
from port.cdm.asset.commercial.schema import COMMERCIAL_SCHEMA  # noqa: E402
from port.cdm.asset.loan.schema import MORTGAGE_SCHEMA  # noqa: E402
from port.cdm.asset.residential.schema import PROPERTY_SCHEMA  # noqa: E402

# Asset label -> CDM schema dict. Mirrors the tool's ASSETS registry; commercial
# loans wrap the residential-loan (RLoan) sections under a legacy "Mortgage" key.
ASSETS = [
    ("Portfolio", PROPERTY_SCHEMA),
    ("Mortgage", MORTGAGE_SCHEMA),
    ("Commercial", COMMERCIAL_SCHEMA),
    ("Commercial Loan", {"Mortgage": MORTGAGE_SCHEMA["RLoan"]}),
]

OUT_PATH = REPO_ROOT / "docs" / "bulkupload" / "cdm_upload_workbook.xlsx"

# styling
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="DDEBF7")
WRAP = Alignment(vertical="top", wrap_text=True)
TOP = Alignment(vertical="top")


def _pretty(seg: str) -> str:
    """A human label for a CamelCase / snake field segment."""
    s = seg.replace("_", " ")
    s = re.sub(r"(?<=[a-z0-9])(?=[A-Z])", " ", s)
    s = re.sub(r"(?<=[A-Z])(?=[A-Z][a-z])", " ", s)
    return s[:1].upper() + s[1:]


def _allowed(spec: dict) -> str:
    """Human 'allowed values' string for a field spec."""
    if spec.get("type") == "menu" and spec.get("options"):
        return ", ".join(str(o) for o in spec["options"])
    if spec.get("input") == "number":
        lo, hi = spec.get("min"), spec.get("max")
        if lo is not None or hi is not None:
            return f"{'' if lo is None else lo} … {'' if hi is None else hi}".strip()
    return ""


def _fields(schema: dict):
    """[(path, section, label, spec)] for every leaf field, in schema order."""
    rows = []
    for path, spec in schema_specs(schema).items():
        section = path.split(".", 1)[0]
        label = _pretty(path.rsplit(".", 1)[-1])
        rows.append((path, section, label, spec))
    return rows


def _autosize(ws, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width


def _dictionary_sheet(wb: Workbook, asset: str, fields) -> None:
    ws = wb.create_sheet(f"{asset} (Dictionary)")
    headers = ["CDM Path", "Section", "Field", "Type", "Allowed values",
               "Description"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment = HEADER_FILL, HEADER_FONT, TOP
    for path, section, label, spec in fields:
        ws.append([path, section, label, spec.get("type", ""),
                   _allowed(spec), spec.get("description") or ""])
        ws.cell(row=ws.max_row, column=6).alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    _autosize(ws, {1: 48, 2: 20, 3: 26, 4: 12, 5: 34, 6: 70})


def _template_sheet(wb: Workbook, asset: str, fields) -> None:
    ws = wb.create_sheet(f"{asset} (Template)")
    # Row 1: the CDM dotted path is the machine header (the upload key).
    for col, (path, section, label, spec) in enumerate(fields, start=1):
        cell = ws.cell(row=1, column=col, value=path)
        cell.fill, cell.font, cell.alignment = HEADER_FILL, HEADER_FONT, WRAP
        bits = [label, f"type: {spec.get('type', '')}"]
        allowed = _allowed(spec)
        if allowed:
            bits.append(f"allowed: {allowed}")
        if spec.get("description"):
            bits.append(str(spec["description"]))
        ws.cell(row=1, column=col).comment = Comment("\n".join(bits), "CDM")
        ws.column_dimensions[get_column_letter(col)].width = 22
        # Dropdown for menu fields (Excel inline-list limit is 255 chars).
        if spec.get("type") == "menu" and spec.get("options"):
            formula = '"' + ",".join(str(o) for o in spec["options"]) + '"'
            if len(formula) <= 255:
                dv = DataValidation(type="list", formula1=formula, allow_blank=True)
                letter = get_column_letter(col)
                dv.add(f"{letter}2:{letter}1000")
                ws.add_data_validation(dv)
    ws.freeze_panes = "A2"


def _overview_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "MKM CDM — asset upload workbook"
    ws["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "Generated from the canonical CDM Python schemas in "
        "src/port/cdm/asset/ via cdm_edit.schema_specs (no hand-maintained "
        "field list — it cannot drift from the schema).",
        "",
        "Each asset class has two sheets:",
        "  • '<Asset> (Dictionary)' — one row per field: CDM path, type, "
        "allowed values, description. The field contract.",
        "  • '<Asset> (Template)'  — fields as columns (CDM path in the header, "
        "rules on hover, dropdowns for menu fields). Fill in one record per row "
        "from row 2 to upload.",
        "",
        "Asset sheets:",
    ]
    for i, (asset, schema) in enumerate(ASSETS):
        notes.append(f"  • {asset} — {len(_fields(schema))} fields")
    for line in notes:
        ws.append([line])
    ws.column_dimensions["A"].width = 100


def build() -> Path:
    wb = Workbook()
    wb.remove(wb.active)  # drop the default empty sheet
    for asset, schema in ASSETS:
        fields = _fields(schema)
        _dictionary_sheet(wb, asset, fields)
        _template_sheet(wb, asset, fields)
    _overview_sheet(wb)
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    return OUT_PATH


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
    for asset, schema in ASSETS:
        print(f"  {asset}: {len(_fields(schema))} fields")
