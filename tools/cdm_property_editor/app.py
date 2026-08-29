# Copyright (c) 2022-2026 MKM Research Labs.
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""
Standalone CDM Asset Review — side tool.

A self-contained Flask app for browsing the asset CDMs as a governance-style
workspace: a top tab bar (Gauges · Properties · Commercials · Mortgage ·
Commercial Loan), a left-hand list of every record in the active tab, and a
per-row review icon that opens the full CDM record in a centered modal with
section tabs. It is schema-driven: each tab's form structure comes straight
from the canonical CDM schema for that asset class.

This is a sandbox tool, deliberately isolated from the production scene:
  * It writes ONLY sandbox copies under
    ``tools/cdm_property_editor/data/<asset>_sandbox.json``.
  * Each sandbox is seeded once from the active catchment's portfolio, read
    through the ``database`` seam (file|pg per MKM_REPO_BACKEND) — not by reading
    ``data/input/<catchment>/`` directly. The real ``data/`` tree is never modified.

Run:
    source .venv/bin/activate
    python tools/cdm_property_editor/app.py
    # then open http://127.0.0.1:5057
"""

from __future__ import annotations

import json
import sys
import uuid
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, render_template, request

# --- Locate the repo and make the CDM schemas importable --------------------
TOOL_DIR = Path(__file__).resolve().parent
REPO_ROOT = TOOL_DIR.parent.parent
SRC_DIR = REPO_ROOT / "src"
# Importing the CDM schemas pulls in the wider `port` package whose __init__
# chain reaches the top-level `config` package at the repo root — so both
# must be importable.
for _p in (str(SRC_DIR), str(REPO_ROOT)):
    if _p not in sys.path:
        sys.path.insert(0, _p)

import price_new  # noqa: E402
from recompute import MODE_DIRS_BY_ASSET, recompute_decomposition  # noqa: E402

# Catchment data is read through the database seam (file|pg per MKM_REPO_BACKEND),
# not by reading data/input/<catchment> directly. WP3.1.
import database  # noqa: E402
from cdm_edit import descriptor_at, schema_specs, validate_value  # noqa: E402
from config import config  # noqa: E402
from database.config_binding import use_configured_backend  # noqa: E402
from lineage.field_usage import AMBER_PREFIXES, EXACT_FIELDS, TIER_META  # noqa: E402
from lineage.field_usage.resolve import classify  # noqa: E402
from port.cdm.asset.commercial.schema import COMMERCIAL_SCHEMA  # noqa: E402
from port.cdm.asset.loan.schema import MORTGAGE_SCHEMA  # noqa: E402
from port.cdm.asset.residential.schema import PROPERTY_SCHEMA  # noqa: E402
from port.cdm.ctpy._schema import COUNTERPARTY_SCHEMA  # noqa: E402
from port.cdm.gauge.schema import GAUGE_SCHEMA  # noqa: E402

# The PRS waterfall is rendered by the main app's own renderer
# (src/static/js/property/phc_basis_waterfall.js) — reused here as a shared
# utility rather than duplicated, so the tool's PRS Waterfall tab is the exact
# same chart as the production basis-explorer right panel.
SHARED_WATERFALL_JS = SRC_DIR / "static" / "js" / "property" / "phc_basis_waterfall.js"

# Bind the data-access backend selected by MKM_REPO_BACKEND (default 'file').
use_configured_backend()
# Active catchment from config (honours MKM_CATCHMENT) — was hardcoded "thames".
CATCHMENT = config.catchment_id
# Every edit persists through the seam under a scratch catchment (a sandbox port_run),
# isolated from the real catchment and from production data/. WP3.2.
SANDBOX_CATCHMENT = f"{CATCHMENT}__cdm_sandbox"
# Still used by the file-based recompute / price_new subprocess path (WP3.2);
# catchment-aware via config rather than a hardcoded thames path.
INPUT_DIR = config.get_input_dir()
GOLDEN_PROPERTY = REPO_ROOT / "tests" / "port" / "cdm" / "golden" / "property.json"
SANDBOX_DIR = TOOL_DIR / "data"
AUDIT_FILE = SANDBOX_DIR / "audit_log.json"
# No authentication yet — every amendment is attributed to a placeholder user
# until real users / sign-in arrive.
AUDIT_USER = "Placeholder User"


# --- Per-asset summary builders (the left-list row + modal header card) ------
def _addr(loc: dict) -> str:
    bits = [loc.get("BuildingNumber"), loc.get("StreetName"),
            loc.get("TownCity"), loc.get("Postcode")]
    return " ".join(str(b) for b in bits if b)


def _type_class(t: str | None) -> str:
    return {"residential": "badge-residential", "commercial": "badge-commercial",
            "industrial": "badge-industrial"}.get(t or "", "badge-na")


def _coords(lat, lon) -> dict:
    """Normalise a lat/lon pair to floats, or None when absent/invalid."""
    try:
        return {"lat": float(lat), "lon": float(lon)}
    except (TypeError, ValueError):
        return {"lat": None, "lon": None}


def _sum_gauge(r: dict) -> dict:
    g = r.get("FloodGauge", {})
    h = g.get("Header", {})
    loc = g.get("Location", {})
    return {"id": h.get("GaugeID"), "sub": h.get("GaugeName") or "—",
            "tag": h.get("CatchmentID") or "gauge", "tagClass": "badge-industrial",
            "value": None,
            **_coords(loc.get("GaugeLatitude"), loc.get("GaugeLongitude"))}


def _sum_property(r: dict) -> dict:
    h = r.get("PropertyHeader", {})
    hd = h.get("Header", {})
    loc = h.get("Location", {})
    return {"id": hd.get("PropertyID") or hd.get("UPRN"), "sub": _addr(loc),
            "tag": hd.get("propertyType"), "tagClass": _type_class(hd.get("propertyType")),
            "value": h.get("Valuation", {}).get("PropertyValue"),
            **_coords(loc.get("LatitudeDegrees"), loc.get("LongitudeDegrees"))}


def _sum_commercial(r: dict) -> dict:
    a = r.get("CommercialAsset", {})
    hd = a.get("Header", {})
    loc = a.get("Location", {})
    ct = a.get("CommercialAttributes", {}).get("CommercialType")
    return {"id": hd.get("PropertyID") or hd.get("UPRN"), "sub": _addr(loc),
            "tag": ct or "commercial", "tagClass": "badge-commercial",
            "value": a.get("Valuation", {}).get("PropertyValue"),
            **_coords(loc.get("LatitudeDegrees"), loc.get("LongitudeDegrees"))}


def _sum_loan(r: dict, root: str = "RLoan") -> dict:
    ln = r.get(root, {})
    h = ln.get("Header", {})
    cur = ln.get("CurrentStatus", {})
    fin = ln.get("FinancialTerms", {})
    sub = ("Property " + h.get("PropertyID")) if h.get("PropertyID") else "—"
    return {"id": h.get("RLoanID") or h.get("MortgageID"), "sub": sub,
            "tag": cur.get("AccountStatus") or "loan", "tagClass": "badge-status",
            "value": cur.get("OutstandingBalance") or fin.get("OriginalLoan")}


def _sum_commercial_loan(r: dict) -> dict:
    s = _sum_loan(r, root="Mortgage")
    meta = r.get("_commercial_meta", {})
    bits = [meta.get("commercial_type"), meta.get("borrower_type")]
    extra = " · ".join(b for b in bits if b)
    if extra:
        s["sub"] = extra
    return s


def _sum_counterparty(r: dict) -> dict:
    party = r.get("CounterpartySet", {}).get("Party", {})
    contact = party.get("ContactInformation", {})
    loc = " ".join(b for b in [contact.get("City"), contact.get("Country")] if b)
    return {"id": party.get("PartyID"), "sub": party.get("PartyName") or "—",
            "tag": loc or "counterparty", "tagClass": "badge-na", "value": None}


# --- Asset registry: drives the top tabs and every endpoint ------------------
# Order here is the top-tab order in the UI.
ASSETS: dict[str, dict] = {
    "gauge": {
        "label": "Gauges", "schema": GAUGE_SCHEMA, "file": "gauge.json",
        "container": "flood_gauges", "summary": _sum_gauge,
    },
    "property": {
        "label": "Properties", "schema": PROPERTY_SCHEMA, "file": "property.json",
        "container": "properties", "summary": _sum_property,
        "golden": GOLDEN_PROPERTY,
    },
    "commercial": {
        "label": "Commercials", "schema": COMMERCIAL_SCHEMA, "file": "commercial.json",
        "container": "commercial_assets", "summary": _sum_commercial,
    },
    "mortgage": {
        "label": "Mortgage", "schema": MORTGAGE_SCHEMA, "file": "loan.json",
        "container": "loans", "summary": _sum_loan,
    },
    "commercial_loan": {
        "label": "Commercial Loan",
        # commercial_loan records wrap the loan under a legacy "Mortgage" key
        # whose inner sections are identical to the RLoan schema.
        "schema": {"Mortgage": MORTGAGE_SCHEMA["RLoan"]},
        "file": "commercial_loan.json", "container": "commercial_loans",
        "summary": _sum_commercial_loan,
    },
    "counterparty": {
        "label": "Counterparty", "schema": COUNTERPARTY_SCHEMA,
        "file": "counterparty.json", "container": "counterparties",
        "summary": _sum_counterparty,
    },
}

# Per-asset hazard-curve files carrying the per-storm damage detail. Keyed by
# the same record id used in the item lists. Only flood-bearing assets appear.
HC_CONFIG: dict[str, dict] = {
    "property": {"file": "propertyhc.json", "container": "property_hazard_curves"},
    "commercial": {"file": "commercialhc.json", "container": "property_hazard_curves"},
}

# Fire / seismic model outputs. Both are commercial-only (keyed by CPROP id) and
# carry per-asset outcome distributions rather than per-event lists. Read via the seam.
_MODEL_GETTERS = {
    "fire": database.get_fire_results,
    "seismic": database.get_seismic_results,
}
# Per-asset portfolio + hazard-curve getters (the seam returns the same document
# shape as the old data/input/<catchment>/<file>.json).
_PORTFOLIO_GETTERS = {
    "gauge": database.get_gauge_portfolio,
    "property": database.get_property_portfolio,
    "commercial": database.get_commercial_portfolio,
    "mortgage": database.get_loan_portfolio,
    "commercial_loan": database.get_commercial_loan_portfolio,
    "counterparty": database.get_counterparty_portfolio,
}
_HC_GETTERS = {
    "property": database.get_property_hazard_curves,
    "commercial": database.get_commercial_hazard_curves,
}
# Sandbox writes go back through the seam (to the scratch catchment), never to disk.
_SAVE_FNS = {
    "gauge": database.save_gauges,
    "property": database.save_properties,
    "commercial": database.save_commercial,
    "mortgage": database.save_loans,
    "commercial_loan": database.save_commercial_loans,
    "counterparty": database.save_counterparties,
}
_HC_SAVE_FNS = {
    "property": database.save_property_hazard_curves,
    "commercial": database.save_commercial_hazard_curves,
}
# Seismic damage states DS0..DS3; DS3 is the collapse state (no_collapse counts
# DS0+DS1+DS2). See src/models/seismic/damage.py.
SEISMIC_DS_LABELS = ["None", "Slight", "Moderate", "Collapse"]

# Lazy caches (read-only reference data; never written back).
_CACHE: dict = {}


def _model_assets(cache_key: str) -> tuple:
    """(model_id, {asset_id: record}) for a fire/seismic model output, via the seam."""
    if cache_key not in _CACHE:
        doc = _MODEL_GETTERS[cache_key](CATCHMENT) or {}
        model = doc.get("metadata", {}).get("model")
        amap = {a.get("asset_id"): a for a in doc.get("assets", [])}
        _CACHE[cache_key] = (model, amap)
    return _CACHE[cache_key]


def _storm_type_index() -> dict:
    """sequence_id -> {type (cluster), intensity}; built once from storm_sequences."""
    if "storm_types" not in _CACHE:
        idx = {}
        doc = database.get_storm_sequences(CATCHMENT) or {}
        for s in doc.get("sequences", []):
            idx[s.get("sequence_id")] = {
                "type": s.get("sequence_type"),
                "intensity": s.get("intensity_category"),
            }
        _CACHE["storm_types"] = idx
    return _CACHE["storm_types"]


def _load_priced(asset: str) -> dict:
    """On-demand-priced new-asset hazard curves, held in the sandbox scratch run
    (``{rid: curve}``), never on disk."""
    if asset not in _HC_GETTERS:
        return {}
    doc = _HC_GETTERS[asset](SANDBOX_CATCHMENT) or {}
    return doc.get(HC_CONFIG[asset]["container"], {})


def _save_priced(asset: str, rid: str, curve: dict) -> None:
    store = _load_priced(asset)
    store[rid] = curve
    _HC_SAVE_FNS[asset](SANDBOX_CATCHMENT, {HC_CONFIG[asset]["container"]: store})


def _hc_record(asset: str, rid: str) -> dict | None:
    # A newly-added asset priced on demand lives in the sandbox store; prefer it.
    priced = _load_priced(asset)
    if rid in priced:
        return priced[rid]
    cfg = HC_CONFIG.get(asset)
    if not cfg:
        return None
    key = f"hc_{asset}"
    if key not in _CACHE:
        doc = _HC_GETTERS[asset](CATCHMENT) or {}
        _CACHE[key] = doc.get(cfg["container"], {})
    return _CACHE[key].get(rid)


app = Flask(__name__)

# The design tokens (coding rule R7). This tool's stylesheet used to carry its own
# ``:root`` block of eleven colours; it now refers to the shared vocabulary in
# ``config.theme``, so the ``:root`` has to arrive from somewhere. Registering the
# platform's theme blueprint serves it at ``/theme.css``, which ``templates/index.html``
# links ahead of its own stylesheet. Same tokens, same values, one source — this tool
# and the console cannot drift apart.
from routes.theme import theme_bp  # noqa: E402

app.register_blueprint(theme_bp)
# Preserve the curated CDM key order (schema + records) instead of
# alphabetising it — the section/field order is meaningful.
app.json.sort_keys = False


# --- Sandbox plumbing (a scratch catchment behind the seam, never data/) -----
def _seed_doc(asset: str) -> dict:
    """The baseline portfolio document used to seed a fresh sandbox, read via the
    seam. Falls back to a golden fixture if the seam has no records for this asset."""
    doc = _PORTFOLIO_GETTERS[asset](CATCHMENT) or {}
    if _records_of(doc, asset):
        return doc
    golden = ASSETS[asset].get("golden")
    if golden and Path(golden).exists():
        return json.loads(Path(golden).read_text())
    raise FileNotFoundError(
        f"No seed data for '{asset}' in catchment '{CATCHMENT}' "
        "(empty portfolio and no golden fixture)."
    )


def _load_doc(asset: str):
    """The asset's sandbox portfolio document, read from the scratch catchment via the
    seam. Seeded once from the real catchment's baseline on first use."""
    doc = _PORTFOLIO_GETTERS[asset](SANDBOX_CATCHMENT)
    if not doc:
        doc = _seed_doc(asset)
        _SAVE_FNS[asset](SANDBOX_CATCHMENT, doc)
    return doc


def _save_doc(asset: str, doc) -> None:
    """Persist the edited document to the sandbox scratch catchment (never data/)."""
    _SAVE_FNS[asset](SANDBOX_CATCHMENT, doc)


def _records_of(doc, asset: str) -> list:
    if isinstance(doc, list):
        return doc
    return doc.get(ASSETS[asset]["container"]) or []


def _records(asset: str) -> list:
    return _records_of(_load_doc(asset), asset)


def _set_nested(rec: dict, path: str, value) -> None:
    """Set ``rec`` at a dotted ``path`` (the record root is the section key)."""
    node = rec
    segs = path.split(".")
    for seg in segs[:-1]:
        node = node.setdefault(seg, {})
    node[segs[-1]] = value


def _get_nested(rec: dict, path: str):
    node = rec
    for seg in path.split("."):
        if not isinstance(node, dict) or seg not in node:
            return None
        node = node[seg]
    return node


def _load_audit() -> list:
    if not AUDIT_FILE.exists():
        return []
    try:
        with open(AUDIT_FILE, "r", encoding="utf-8") as fh:
            return json.load(fh)
    except (OSError, json.JSONDecodeError):
        return []


def _append_audit(entries: list) -> None:
    if not entries:
        return
    SANDBOX_DIR.mkdir(parents=True, exist_ok=True)
    log = _load_audit()
    log.extend(entries)
    with open(AUDIT_FILE, "w", encoding="utf-8") as fh:
        json.dump(log, fh, indent=2)


def _record_id(asset: str, rec: dict) -> str:
    return ASSETS[asset]["summary"](rec).get("id") or ""


# --- Routes -----------------------------------------------------------------
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/assets")
def api_assets():
    """Top-tab list: key + label, in display order."""
    return jsonify([{"key": k, "label": v["label"]} for k, v in ASSETS.items()])


@app.route("/api/catchment-info")
def api_catchment_info():
    """Active catchment + its map centre — used as the lat/lon ballpark on the
    add form so a transposed pair is obvious (no hardcoded coordinates)."""
    lat = lon = None
    try:
        from config.visual import get_map_center
        lat, lon = get_map_center()
    except Exception:
        pass
    return jsonify({"catchment": CATCHMENT, "lat": lat, "lon": lon})


@app.route("/api/lineage")
def api_lineage():
    """Field-usage lineage: RED/AMBER/GREEN tiers + downstream chains.

    Served whole (small) so the front end can colour every field box and open
    the per-field lineage popup without a round-trip per field. Matching mirrors
    lineage.field_usage.resolve: exact path, then AMBER prefix, else GREEN.
    """
    return jsonify({
        "tiers": TIER_META,
        "exact": EXACT_FIELDS,
        "amberPrefixes": [{"prefix": p, "entry": e} for p, e in AMBER_PREFIXES],
    })


@app.route("/api/<asset>/schema")
def api_schema(asset: str):
    cfg = ASSETS.get(asset)
    if not cfg:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    return jsonify({"sections": list(cfg["schema"].keys()), "schema": cfg["schema"]})


@app.route("/api/<asset>/edit-spec")
def api_edit_spec(asset: str):
    """Per-field editor specs (input widget, menu options, numeric bounds).

    Keyed by dotted path from the record root, so the editor builds the right
    widget for each field and the same bounds the server enforces on commit.
    """
    cfg = ASSETS.get(asset)
    if not cfg:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    return jsonify(schema_specs(cfg["schema"]))


@app.route("/api/<asset>/items")
def api_items(asset: str):
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    cfg = ASSETS[asset]
    return jsonify([cfg["summary"](r) for r in _records(asset)])


@app.route("/api/<asset>/items/<rid>")
def api_item(asset: str, rid: str):
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    for rec in _records(asset):
        if _record_id(asset, rec) == rid:
            return jsonify(rec)
    return jsonify({"error": f"{asset} '{rid}' not found"}), 404


def _baseline_value(asset: str, rid: str, path: str):
    """The field value in the read-only data/ baseline (pre-sandbox edits).

    The recompute applies the change relative to the baseline (which the stored
    timeseries reflects), so before/after is always original-vs-current.
    """
    try:
        doc = _seed_doc(asset)
    except FileNotFoundError:
        return None
    rec = next((r for r in _records_of(doc, asset) if _record_id(asset, r) == rid), None)
    return _get_nested(rec, path) if rec else None


def _num_storms(asset: str) -> int | None:
    if asset not in _HC_GETTERS:
        return None
    doc = _HC_GETTERS[asset](CATCHMENT) or {}
    return doc.get("metadata", {}).get("num_storms")


def _maybe_recompute(asset: str, rid: str, target: dict, changed: dict) -> dict | None:
    """RED-gated before/after PRS recompute for a committed edit.

    Returns ``None`` when no RED field changed. For a RED change, returns the
    instant before/after spread decomposition (shortcut), or a not-supported
    note (gauge-threshold / ground edits → "needs a full re-run", deferred).
    Property and commercial assets are priced; others have no hazard curve.
    """
    if asset not in MODE_DIRS_BY_ASSET:
        return None
    for path in changed:
        if classify(path) != "RED":
            continue
        before = (_hc_record(asset, rid) or {}).get("spread_decomposition")
        num = _num_storms(asset)
        if not before or not num:
            return {"supported": False, "tier": "RED", "field": path,
                    "reason": "No hazard curve for this asset yet."}
        after = recompute_decomposition(
            rid, path, _baseline_value(asset, rid, path), _get_nested(target, path),
            asset=asset, catchment=CATCHMENT, before_decomp=before, num_storms=num)
        if after is None:
            return {"supported": False, "tier": classify(path), "field": path,
                    "reason": "This edit needs a full portfolio re-run, "
                              "not an instant recompute."}
        return {"supported": True, "tier": "RED", "field": path,
                "before": before, "after": after}
    return None


# Assets the tool can CREATE (zero-portfolio / add). Maps the minted id + the
# catchment tag onto the right CDM paths so the record joins the thames pool.
NEW_ASSET: dict[str, dict] = {
    "property": {"id_path": "PropertyHeader.Header.PropertyID", "prefix": "PROP-",
                 "catchment_path": "PropertyHeader.Header.CatchmentID"},
    "commercial": {"id_path": "CommercialAsset.Header.PropertyID", "prefix": "CPROP-",
                   "catchment_path": "CommercialAsset.Header.CatchmentID"},
    "mortgage": {"id_path": "RLoan.Header.RLoanID", "prefix": "RLOAN-",
                 "catchment_path": "RLoan.Header.CatchmentID"},
    "commercial_loan": {"id_path": "Mortgage.Header.RLoanID", "prefix": "CLOAN-",
                        "catchment_path": "Mortgage.Header.CatchmentID"},
}


@app.route("/api/<asset>/items", methods=["POST"])
def api_create(asset: str):
    """Create a new asset record in the SANDBOX (never data/) — the add path.

    Body: ``{"changes": {dotted_path: value, ...}}`` (same shape as the editor).
    Mints the id, tags the catchment, validates every field, appends and audits
    the creation. Returns the new id so the UI can open it.
    """
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    cfg = NEW_ASSET.get(asset)
    if cfg is None:
        return jsonify({"error": f"Creating '{asset}' records is not supported yet."}), 400

    schema = ASSETS[asset]["schema"]
    changes = (request.get_json(silent=True) or {}).get("changes", {})
    coerced, errors = {}, {}
    for path, raw in changes.items():
        if raw in (None, ""):
            continue  # skip blanks — a new record need not fill every field
        descriptor = descriptor_at(schema, path)
        if descriptor is None:
            errors[path] = "not an editable schema field"
            continue
        ok, value, err = validate_value(path, descriptor, raw)
        if ok:
            coerced[path] = value
        else:
            errors[path] = err
    if errors:
        return jsonify({"status": "error", "errors": errors}), 400

    new_id = cfg["prefix"] + uuid.uuid4().hex[:8]
    rec: dict = {}
    _set_nested(rec, cfg["id_path"], new_id)
    _set_nested(rec, cfg["catchment_path"], CATCHMENT)
    for path, value in coerced.items():
        _set_nested(rec, path, value)

    doc = _load_doc(asset)
    if isinstance(doc, list):
        doc.append(rec)
    else:
        doc.setdefault(ASSETS[asset]["container"], []).append(rec)
    _save_doc(asset, doc)
    _append_audit([{
        "timestamp": datetime.now().isoformat(timespec="seconds"), "user": AUDIT_USER,
        "asset": asset, "asset_label": ASSETS[asset]["label"], "record_id": new_id,
        "field": "(record)", "field_label": "Created record", "old": None, "new": new_id,
    }])
    return jsonify({"status": "success", "id": new_id, "record": rec})


# The workbook Template sheet that maps to each creatable asset (header row =
# CDM dotted paths). Matches tools/cdm_property_editor/cdm_workbook.py.
ASSET_TEMPLATE_SHEET = {
    "property": "Portfolio (Template)",
    "commercial": "Commercial (Template)",
    "mortgage": "Mortgage (Template)",
    "commercial_loan": "Commercial Loan (Template)",
}


@app.route("/api/<asset>/upload", methods=["POST"])
def api_upload(asset: str):
    """Bulk-add records from an uploaded .xlsx (the CDM upload workbook).

    Reads the asset's Template sheet — header row = CDM dotted paths — validates
    every cell against the schema, mints id + catchment, appends the valid rows to
    the sandbox and reports the rejected rows with per-cell reasons.
    """
    cfg = NEW_ASSET.get(asset)
    if cfg is None:
        return jsonify({"error": f"Uploading '{asset}' is not supported."}), 400
    f = request.files.get("file")
    if f is None:
        return jsonify({"error": "No file uploaded (form field 'file')."}), 400

    import io

    import openpyxl
    try:
        wb = openpyxl.load_workbook(io.BytesIO(f.read()), data_only=True, read_only=True)
    except Exception as exc:
        return jsonify({"error": f"Could not read workbook: {exc}"}), 400
    want = ASSET_TEMPLATE_SHEET.get(asset)
    sheet = wb[want] if want in wb.sheetnames else wb[wb.sheetnames[0]]

    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) < 2:
        return jsonify({"created": 0, "ids": [], "rejected": [],
                        "note": f"No data rows in sheet '{sheet.title}'."})
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]

    schema = ASSETS[asset]["schema"]
    doc = _load_doc(asset)
    container = ASSETS[asset]["container"]
    ts = datetime.now().isoformat(timespec="seconds")
    created, rejected, audit = [], [], []
    for ri, row in enumerate(rows[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in row):
            continue  # blank row
        coerced, errors = {}, {}
        for header, raw in zip(headers, row):
            if not header or raw is None or str(raw).strip() == "":
                continue
            descriptor = descriptor_at(schema, header)
            if descriptor is None:
                errors[header] = "unknown field"
                continue
            ok, value, err = validate_value(header, descriptor, raw)
            if ok:
                coerced[header] = value
            else:
                errors[header] = err
        if errors:
            rejected.append({"row": ri, "errors": errors})
            continue
        new_id = coerced.get(cfg["id_path"]) or (cfg["prefix"] + uuid.uuid4().hex[:8])
        rec: dict = {}
        _set_nested(rec, cfg["id_path"], new_id)
        _set_nested(rec, cfg["catchment_path"], CATCHMENT)
        for path, value in coerced.items():
            _set_nested(rec, path, value)
        doc.setdefault(container, []).append(rec)
        created.append(new_id)
        audit.append({"timestamp": ts, "user": AUDIT_USER, "asset": asset,
                      "asset_label": ASSETS[asset]["label"], "record_id": new_id,
                      "field": "(record)", "field_label": "Uploaded record",
                      "old": None, "new": new_id})
    if created:
        _save_doc(asset, doc)
        _append_audit(audit)
    return jsonify({"created": len(created), "ids": created,
                    "rejected": rejected, "sheet": sheet.title})


@app.route("/api/<asset>/items/<rid>/price", methods=["POST"])
def api_price(asset: str, rid: str):
    """Price a newly-added asset that has no timeseries yet (Workstream B / B4).

    Runs the real batch generators over a scoped 1-asset workspace (subprocess,
    never touches data/), stores the resulting hazard-curve in the sandbox so the
    PRS Waterfall renders it. Slow (~tens of seconds) — the UI calls it async.
    """
    if asset not in price_new._ASSET:
        return jsonify({"ok": False, "error": f"Pricing '{asset}' is not supported."}), 400
    doc = _load_doc(asset)
    rec = next((r for r in _records_of(doc, asset) if _record_id(asset, r) == rid), None)
    if rec is None:
        return jsonify({"ok": False, "error": f"{asset} '{rid}' not found"}), 404
    res = price_new.price_asset(asset, rid, rec, input_dir=INPUT_DIR,
                                src_dir=SRC_DIR, repo_root=REPO_ROOT)
    if not res["ok"]:
        return jsonify({"ok": False, "error": res["error"]}), 200
    _save_priced(asset, rid, res["curve"])
    sd = res["curve"].get("spread_decomposition", {}) or {}
    return jsonify({
        "ok": True,
        "spread_decomposition": sd,
        "property_spread_bps": sd.get("property_spread_bps") or 0.0,
        "flood_count": res["curve"].get("flood_count"),
        "flood_zone": res["curve"].get("flood_zone"),
    })


@app.route("/api/<asset>/items/<rid>", methods=["PUT"])
def api_update(asset: str, rid: str):
    """Commit edited field values to the SANDBOX copy (never to data/).

    Body: ``{"changes": {dotted_path: new_value, ...}}``. Each change is
    validated/coerced against its schema descriptor (menu membership, numeric
    bounds, type). On any error nothing is written and the per-field errors are
    returned with 400.
    """
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    schema = ASSETS[asset]["schema"]
    changes = (request.get_json(silent=True) or {}).get("changes", {})
    if not isinstance(changes, dict):
        return jsonify({"error": "changes must be an object"}), 400

    coerced, errors = {}, {}
    for path, raw in changes.items():
        descriptor = descriptor_at(schema, path)
        if descriptor is None:
            errors[path] = "not an editable schema field"
            continue
        ok, value, err = validate_value(path, descriptor, raw)
        if ok:
            coerced[path] = value
        else:
            errors[path] = err
    if errors:
        return jsonify({"status": "error", "errors": errors}), 400

    doc = _load_doc(asset)
    target = next((r for r in _records_of(doc, asset) if _record_id(asset, r) == rid), None)
    if target is None:
        return jsonify({"error": f"{asset} '{rid}' not found"}), 404

    ts = datetime.now().isoformat(timespec="seconds")
    audit = []
    for path, value in coerced.items():
        old = _get_nested(target, path)
        if old == value:
            continue  # no-op change — not audited
        _set_nested(target, path, value)
        audit.append({
            "timestamp": ts, "user": AUDIT_USER,
            "asset": asset, "asset_label": ASSETS[asset]["label"],
            "record_id": rid, "field": path,
            "field_label": path.rsplit(".", 1)[-1],
            "old": old, "new": value,
        })
    _save_doc(asset, doc)
    _append_audit(audit)
    recompute = _maybe_recompute(asset, rid, target, coerced)
    return jsonify({"status": "success", "updated": len(audit),
                    "record": target, "recompute": recompute})


@app.route("/api/audit")
def api_audit():
    """The amendment audit trail (newest first)."""
    return jsonify(list(reversed(_load_audit())))


def _flood_payload(asset: str, rid: str) -> dict:
    """Severe flood events: storm, cluster type, damage %. Property/commercial."""
    if asset not in HC_CONFIG:
        return {"supported": False, "reason": "No flood model for this asset."}
    rec = _hc_record(asset, rid)
    if rec is None:
        return {"supported": True, "count": 0, "events": []}
    idx = _storm_type_index()
    events = []
    for d in rec.get("storm_details", []):
        if not (d.get("flooded") or d.get("exceeded_severe") or d.get("damage_ratio", 0) > 0):
            continue
        t = idx.get(d.get("storm_id"), {})
        events.append({
            "storm": d.get("storm_id"),
            "type": t.get("type") or "—",
            "intensity": t.get("intensity"),
            "damage_pct": round(d.get("damage_ratio", 0) * 100, 2),
            "depth_m": round(d.get("flood_depth_m", 0) or 0, 2),
            "peak_m": round(d.get("gauge_peak_m", 0) or 0, 2),
            "severe": bool(d.get("exceeded_severe")),
        })
    events.sort(key=lambda e: (-e["damage_pct"], 0 if e["severe"] else 1))
    return {"supported": True, "count": len(events),
            "flood_zone": rec.get("flood_zone"), "events": events[:12]}


def _wind_payload(asset: str, rid: str) -> dict:
    """Wind/typhoon events. The thames win files are zero-event placeholders
    (the typhoon ensemble was not run), so there is nothing real to show yet."""
    if asset not in HC_CONFIG:
        return {"supported": False, "reason": "No wind model for this asset."}
    return {"supported": True, "count": 0, "events": [],
            "note": "Typhoon/wind ensemble not run for this catchment — "
                    "zero-event placeholder. Run `port --typhoon` to populate."}


def _fire_payload(asset: str, rid: str) -> dict:
    """Fire model outcome distribution (commercial assets only)."""
    model, amap = _model_assets("fire")
    a = amap.get(rid)
    if not a:
        return {"supported": False,
                "reason": "Fire model (MKM-FIRE-001) covers commercial assets only."}
    return {
        "supported": True, "model": model, "n_sim": a.get("n_sim"),
        "lambda_annual": a.get("lambda_annual"),
        "n_fires": a.get("n_fires"),
        "outcomes": [
            {"label": "Contained", "count": a.get("n_contained", 0), "cls": "ok"},
            {"label": "Partial loss", "count": a.get("n_partial", 0), "cls": "warn"},
            {"label": "Total loss", "count": a.get("n_total", 0), "cls": "bad"},
            {"label": "Point of no return", "count": a.get("n_point_of_no_return", 0), "cls": "bad"},
        ],
        "stats": [
            {"label": "Loss frequency (annual)", "value": a.get("loss_frequency")},
            {"label": "Partial-loss freq", "value": a.get("partial_loss_frequency")},
            {"label": "Total-loss freq", "value": a.get("total_loss_frequency")},
            {"label": "Containment rate", "value": a.get("containment_rate"), "pct": True},
        ],
    }


def _seismic_payload(asset: str, rid: str) -> dict:
    """Seismic model damage-state distribution (commercial assets only)."""
    model, amap = _model_assets("seismic")
    a = amap.get(rid)
    if not a:
        return {"supported": False,
                "reason": "Seismic model covers commercial assets only."}
    ds = a.get("damage_state_counts", {}) or {}
    classes = ["ok", "warn", "warn", "bad"]
    return {
        "supported": True, "model": model, "n_sim": a.get("n_sim"),
        "n_events": a.get("n_events"),
        "site": [
            {"label": "Hazard zone", "value": a.get("zone")},
            {"label": "Site class", "value": a.get("site_class")},
            {"label": "Construction", "value": a.get("construction_type")},
            {"label": "BRI rating", "value": a.get("rating")},
        ],
        "damage_states": [
            {"label": SEISMIC_DS_LABELS[i], "count": ds.get(str(i), 0), "cls": classes[i]}
            for i in range(4)
        ],
        "stats": [
            {"label": "No-collapse rate", "value": a.get("no_collapse_rate"), "pct": True},
            {"label": "Loss frequency (annual)", "value": a.get("loss_frequency")},
            {"label": "PML 1-in-475", "value": a.get("pml_475"), "pct": True},
            {"label": "PML 1-in-2475", "value": a.get("pml_2475"), "pct": True},
        ],
    }


@app.route("/api/<asset>/items/<rid>/perils")
def api_perils(asset: str, rid: str):
    """All four perils for one asset, each with a `supported` flag."""
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    return jsonify({
        "flood": _flood_payload(asset, rid),
        "wind": _wind_payload(asset, rid),
        "fire": _fire_payload(asset, rid),
        "seismic": _seismic_payload(asset, rid),
    })


@app.route("/api/<asset>/items/<rid>/waterfall")
def api_waterfall(asset: str, rid: str):
    """PRS spread waterfall for one asset, from its hazard-curve decomposition.

    Mirrors the main app's basis waterfall (Gauge -> SHE elevation -> SHD
    distance -> Property -> BRI resilient), each bar the spread under that
    stage's adjustment plus its effect vs the gauge spread. Property/commercial
    only; other assets return supported=False.
    """
    if asset not in ASSETS:
        return jsonify({"error": f"Unknown asset '{asset}'"}), 404
    if asset not in HC_CONFIG:
        return jsonify({"supported": False,
                        "reason": "No PRS waterfall for this asset class."})
    rec = _hc_record(asset, rid)
    if rec is None:
        return jsonify({"supported": True, "spread_decomposition": None,
                        "can_price": asset in price_new._ASSET,
                        "note": "Not priced yet — this asset has no hazard curve."})
    sd = rec.get("spread_decomposition", {}) or {}
    # Return the raw decomposition; the shared phc_basis_waterfall.js renderer
    # builds the bars (Gauge -> SHE -> SHD -> Property -> BRI) from it, exactly
    # as the main-app basis panel does.
    return jsonify({
        "supported": True,
        "flood_zone": rec.get("flood_zone"),
        "property_spread_bps": sd.get("property_spread_bps") or 0.0,
        "spread_decomposition": sd,
    })


@app.route("/shared/phc_basis_waterfall.js")
def shared_waterfall_js():
    """Serve the main app's spread-waterfall renderer verbatim (shared utility)."""
    return SHARED_WATERFALL_JS.read_text(encoding="utf-8"), 200, \
        {"Content-Type": "application/javascript"}


if __name__ == "__main__":
    print(f"CDM Asset Review — sandbox dir: {SANDBOX_DIR}")
    print(f"Assets: {', '.join(ASSETS)}  (catchment: {CATCHMENT})")
    print("Open http://127.0.0.1:5057")
    app.run(host="127.0.0.1", port=5057, debug=True)
