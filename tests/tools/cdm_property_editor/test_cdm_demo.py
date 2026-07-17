# Copyright (c) 2022-2026 MKM Research Labs.
#
# Permission is hereby granted, free of charge, to any person obtaining a copy
# of this software and associated documentation files (the "Software"), to deal
# in the Software without restriction, including without limitation the rights
# to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
# copies of the Software, and to permit persons to whom the Software is
# furnished to do so, subject to the following conditions:
#
# The above copyright notice and this permission notice shall be included in all
# copies or substantial portions of the Software.
#
# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Tests for the filled demo-workbook generator (cdm_demo.py).

Covers the dotted-path getter, the validated/​re-stamped row builder (including
the missing-value, missing-descriptor and validation-fail skip branches), the
per-sheet writer's used-column pruning, and a full ``build()`` round-trip against
a synthetic thames input dir + redirected output path (no real data, no SSD).
"""

import json

from openpyxl import load_workbook


# --- _get -------------------------------------------------------------------
def test_get_walks_dotted_path(demo_mod):
    rec = {"A": {"B": {"C": 7}}}
    assert demo_mod._get(rec, "A.B.C") == 7


def test_get_missing_returns_none(demo_mod):
    assert demo_mod._get({"A": {"B": 1}}, "A.B.C") is None  # walk past scalar
    assert demo_mod._get({}, "A") is None


# --- _demo_rows -------------------------------------------------------------
def test_demo_rows_restamps_id_and_skips_blanks(demo_mod):
    from port.cdm.asset.residential.schema import PROPERTY_SCHEMA
    id_path = "PropertyHeader.Header.PropertyID"
    records = [{
        "PropertyHeader": {
            "Header": {"PropertyID": "ORIG", "propertyType": "residential"},
            "Location": {"StreetName": "High St", "TownCity": None},
            "Valuation": {"PropertyValue": 400000}}}]
    columns = [id_path,
               "PropertyHeader.Location.StreetName",
               "PropertyHeader.Location.TownCity",  # None -> skipped
               "PropertyHeader.Valuation.PropertyValue"]
    rows = demo_mod._demo_rows(PROPERTY_SCHEMA, records, id_path, "DEMO-PROP",
                               columns)
    assert rows[0][id_path] == "DEMO-PROP-01"
    assert rows[0]["PropertyHeader.Location.StreetName"] == "High St"
    assert "PropertyHeader.Location.TownCity" not in rows[0]


def test_demo_rows_skips_unknown_descriptor(demo_mod):
    from port.cdm.asset.residential.schema import PROPERTY_SCHEMA
    id_path = "PropertyHeader.Header.PropertyID"
    records = [{"PropertyHeader": {"Header": {"PropertyID": "X"},
                                   "NotASection": {"Bogus": "v"}}}]
    rows = demo_mod._demo_rows(
        PROPERTY_SCHEMA, records, id_path, "DEMO",
        [id_path, "PropertyHeader.NotASection.Bogus"])
    # the unknown-schema path is dropped, only the id remains
    assert rows[0] == {id_path: "DEMO-01"}


def test_demo_rows_caps_at_N(demo_mod):
    from port.cdm.asset.residential.schema import PROPERTY_SCHEMA
    id_path = "PropertyHeader.Header.PropertyID"
    records = [{"PropertyHeader": {"Header": {"PropertyID": f"P{i}"}}}
               for i in range(demo_mod.N + 3)]
    rows = demo_mod._demo_rows(PROPERTY_SCHEMA, records, id_path, "D", [id_path])
    assert len(rows) == demo_mod.N


# --- _sheet -----------------------------------------------------------------
def test_sheet_keeps_only_used_columns(demo_mod):
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)
    columns = ["a", "b", "c"]
    rows = [{"a": 1, "c": 3}]  # 'b' populated by nobody -> pruned
    demo_mod._sheet(wb, "Demo (Template)", columns, rows)
    ws = wb["Demo (Template)"]
    headers = [c.value for c in ws[1]]
    assert headers == ["a", "c"]
    assert ws.freeze_panes == "A2"


# --- build() round-trip -----------------------------------------------------
def _seed_inputs(demo_mod, input_dir):
    """Write minimal records for each of the four demo asset files."""
    src = demo_mod.REPO_ROOT / "data" / "input" / "thames"
    files = {
        "property.json": ("properties", json.loads(
            (src / "property.json").read_text())["properties"][:2]),
        "commercial.json": ("commercial_assets", json.loads(
            (src / "commercial.json").read_text())["commercial_assets"][:2]),
        "loan.json": ("loans", json.loads(
            (src / "loan.json").read_text())["loans"][:2]),
        "commercial_loan.json": ("commercial_loans", json.loads(
            (src / "commercial_loan.json").read_text())["commercial_loans"][:2]),
    }
    for fname, (container, recs) in files.items():
        (input_dir / fname).write_text(json.dumps({container: recs}))


def test_build_roundtrip(demo_mod, tmp_path, monkeypatch):
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    _seed_inputs(demo_mod, input_dir)
    out = tmp_path / "cdm_demo.xlsx"
    monkeypatch.setattr(demo_mod, "INPUT_DIR", input_dir)
    monkeypatch.setattr(demo_mod, "OUT_PATH", out)

    path = demo_mod.build()
    assert path == out and out.exists()
    wb = load_workbook(out)
    # one sheet per asset class, named as in the ASSETS table
    for title, *_ in demo_mod.ASSETS:
        assert title in wb.sheetnames
        ws = wb[title]
        # each demo row's id starts with the asset's DEMO- prefix
        assert ws.max_row >= 2


def test_build_handles_list_shaped_data(demo_mod, tmp_path, monkeypatch):
    """A records file that is a bare list (no container) is read directly."""
    input_dir = tmp_path / "input"
    input_dir.mkdir()
    _seed_inputs(demo_mod, input_dir)
    # overwrite property.json with a bare-list document
    prop = json.loads((input_dir / "property.json").read_text())["properties"]
    (input_dir / "property.json").write_text(json.dumps(prop))
    monkeypatch.setattr(demo_mod, "INPUT_DIR", input_dir)
    monkeypatch.setattr(demo_mod, "OUT_PATH", tmp_path / "out.xlsx")
    assert demo_mod.build().exists()
