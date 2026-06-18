# Copyright (c) 2022-2026 MKM Research Labs. All rights reserved.

# This software is licensed by MKM Research Labs for non-commercial 
# research and educational use only. Any commercial use, including 
# but not limited to use in or for products or services offered for sale, 
# internal business operations intended for commercial advantage, or
# research and development conducted for a commercial entity, is expressly
# prohibited unless separately authorized in writing by MKM Research Labs.

# Use, reproduction, distribution, or modification of this code is subject to the
# terms and conditions of the license agreement provided with this software.

# THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
# IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
# FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
# AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
# LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
# OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
# SOFTWARE.

"""Tests for the CDM upload-workbook generator (cdm_workbook.py).

Covers the pure label/allowed-values helpers, the field flattener, the
per-sheet builders (including the menu-dropdown add vs >255-char skip
branches), and a full ``build()`` round-trip loaded back with openpyxl.
"""

from openpyxl import Workbook, load_workbook


# --- pure helpers -----------------------------------------------------------
def test_pretty_snake_and_camel(workbook_mod):
    assert workbook_mod._pretty("flood_zone") == "Flood zone"
    assert workbook_mod._pretty("CamelCase") == "Camel Case"
    assert workbook_mod._pretty("PropertyID") == "Property ID"


def test_allowed_menu(workbook_mod):
    assert workbook_mod._allowed(
        {"type": "menu", "options": ["a", "b", "c"]}) == "a, b, c"


def test_allowed_number_bounds(workbook_mod):
    assert workbook_mod._allowed(
        {"input": "number", "min": 0, "max": 10}) == "0 … 10"
    assert workbook_mod._allowed(
        {"input": "number", "min": 0, "max": None}) == "0 …"
    assert workbook_mod._allowed(
        {"input": "number", "min": None, "max": 5}) == "… 5"


def test_allowed_none(workbook_mod):
    assert workbook_mod._allowed({"type": "text"}) == ""
    assert workbook_mod._allowed({"input": "number"}) == ""  # no bounds


def test_fields_flattens_schema(workbook_mod):
    from port.cdm.asset.residential.schema import PROPERTY_SCHEMA
    rows = workbook_mod._fields(PROPERTY_SCHEMA)
    assert rows and all(len(r) == 4 for r in rows)
    paths = [r[0] for r in rows]
    assert "PropertyHeader.Valuation.PropertyValue" in paths
    # section is the first dotted segment
    assert all(r[1] == r[0].split(".", 1)[0] for r in rows)


# --- sheet builders ---------------------------------------------------------
def test_dictionary_sheet(workbook_mod):
    wb = Workbook()
    fields = [("Sec.Field", "Sec", "Field", {"type": "text",
              "description": "d"})]
    workbook_mod._dictionary_sheet(wb, "Demo", fields)
    ws = wb["Demo (Dictionary)"]
    assert ws.cell(row=1, column=1).value == "CDM Path"
    assert ws.cell(row=2, column=1).value == "Sec.Field"


def test_template_sheet_dropdown_add_vs_skip(workbook_mod):
    wb = Workbook()
    long_opts = [f"option_value_{i}" for i in range(40)]  # joined > 255 chars
    fields = [
        ("Sec.Short", "Sec", "Short", {"type": "menu", "options": ["a", "b"]}),
        ("Sec.Long", "Sec", "Long", {"type": "menu", "options": long_opts}),
        ("Sec.Plain", "Sec", "Plain", {"type": "text"}),
    ]
    workbook_mod._template_sheet(wb, "Demo", fields)
    ws = wb["Demo (Template)"]
    # path is the machine header in row 1
    assert ws.cell(row=1, column=1).value == "Sec.Short"
    # only the short menu gets a dropdown; the long one is skipped
    assert len(ws.data_validations.dataValidation) == 1


def test_build_roundtrip(workbook_mod, tmp_path, monkeypatch):
    out = tmp_path / "wb.xlsx"
    monkeypatch.setattr(workbook_mod, "OUT_PATH", out)
    path = workbook_mod.build()
    assert path == out and out.exists()
    wb = load_workbook(out)
    # Overview is inserted first, then two sheets per asset class
    assert wb.sheetnames[0] == "Overview"
    for asset, _schema in workbook_mod.ASSETS:
        assert f"{asset} (Dictionary)" in wb.sheetnames
        assert f"{asset} (Template)" in wb.sheetnames
